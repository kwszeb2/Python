import urllib
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill

# website scraped
url = "https://en.wikipedia.org/wiki/List_of_national_parks"

# open connection and download the webpage
uClient = uReq(url)

# uploads content into a variable
page_html = uClient.read()
uClient.close()

# html parser
page_soup = soup(page_html, "html.parser")

# Get the headings and data
table = page_soup.find('table' , { 'class' : "wikitable sortable" })
rows = table.find_all('tr')

# creates excel sheet
book = Workbook()
ws = book.active

# Alignment, Font and Fill for the Headings

ws['A1'].alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True ,wrapText=True)
ws['B1'].alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, wrapText=True)
ws['C1'].alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, wrapText=True)
ws['D1'].alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, wrapText=True)
ws['E1'].alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, wrapText=True)

ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20

ws['A1'] = 'Country'
ws['B1'] = 'Oldest year'
ws['C1'] = 'Number of parks'
ws['D1'] = 'Total area km^2'
ws['E1'] = 'Percentage of country area'

ft = Font(name="Arial", size="14", bold=True)

a = ws['A1']
b = ws['B1']
c = ws['C1']
d = ws['D1']
e = ws['E1']

a.fill = PatternFill("solid", fgColor="00CCFFCC")
b.fill = PatternFill("solid", fgColor="00CCFFCC")
c.fill = PatternFill("solid", fgColor="00CCFFCC")
d.fill = PatternFill("solid", fgColor="00CCFFCC")
e.fill = PatternFill("solid", fgColor="00CCFFCC")

a.font = ft
b.font = ft
c.font = ft
d.font = ft
e.font = ft

i = 0

for row in rows:
    cell = row.find_all('td')
    
    for j in range(len(cell)):
        ws.cell(row=i+2, column=1).value = cell[0].find('a').text # country

        ws.cell(row=i+2, column=2).value = cell[1].text # year
        ws.cell(row=i+2, column=2).alignment = Alignment(horizontal='center')

        ws.cell(row=i+2, column=3).value = cell[2].text # number of parks
        ws.cell(row=i+2, column=3).alignment = Alignment(horizontal='center')

        ws.cell(row=i+2, column=4).value = cell[3].text # total area
        ws.cell(row=i+2, column=4).alignment = Alignment(horizontal='center')

        ws.cell(row=i+2, column=5).value = cell[4].text # % of country area
        ws.cell(row=i+2, column=5).alignment = Alignment(horizontal='center')
    i = i + 1    
        
book.save('National_parks.xlsx')
book.close()